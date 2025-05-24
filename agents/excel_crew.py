"""
Excel Generator Crew - Multi-Agent System

This module implements a comprehensive Excel generation system using CrewAI,
with specialized agents for different aspects of Excel creation.
"""

import os
import json
import pandas as pd
from typing import Dict, Any, List, Optional
from datetime import datetime
import tempfile

# CrewAI imports
from crewai import Agent, Task, Crew, Process
from crewai.tools import BaseTool
from langchain.llms.base import LLM
from langchain.schema import Generation, LLMResult
from langchain.callbacks.manager import CallbackManagerForLLMRun

# Import existing Excel generation capabilities
from excel_generator import (
    generate_excel_from_prompt,
    create_excel_file,
    parse_gemini_response
)

# Custom LLM wrapper for Gemini to work with CrewAI
class GeminiLLM(LLM):
    """Custom LLM wrapper for Google Gemini to work with CrewAI"""
    
    model: str = "gemini-2.0-flash"
    
    def __init__(self, model_name: str = "gemini-2.0-flash", **kwargs):
        super().__init__(model=model_name, **kwargs)
        import google.generativeai as genai
        # Use object.__setattr__ to bypass Pydantic validation
        object.__setattr__(self, 'genai', genai)
        object.__setattr__(self, 'genai_model', genai.GenerativeModel(model_name))
    
    @property
    def _llm_type(self) -> str:
        return "gemini"
    
    def _call(
        self,
        prompt: str,
        stop: Optional[List[str]] = None,
        run_manager: Optional[CallbackManagerForLLMRun] = None,
        **kwargs: Any,
    ) -> str:
        try:
            response = self.genai_model.generate_content(prompt)
            return response.text
        except Exception as e:
            return f"Error generating response: {str(e)}"
    
    def _generate(
        self,
        prompts: List[str],
        stop: Optional[List[str]] = None,
        run_manager: Optional[CallbackManagerForLLMRun] = None,        **kwargs: Any,
    ) -> LLMResult:
        generations = []
        for prompt in prompts:
            text = self._call(prompt, stop, run_manager, **kwargs)
            generations.append([Generation(text=text)])
        return LLMResult(generations=generations)


# Custom Tools for Excel Generation
class DataAnalysisTool(BaseTool):
    """Tool for analyzing data requirements and structure"""
    
    name: str = "data_analysis_tool"
    description: str = "Analyzes natural language requests to determine data structure, types, and relationships needed for Excel generation"
    
    def _run(self, query: str) -> str:
        """Analyze the user's query to extract data requirements"""
        try:
            # Use the existing Gemini model for analysis
            import google.generativeai as genai
            model = genai.GenerativeModel('gemini-2.0-flash')
            
            analysis_prompt = f"""
            Analyze this request for Excel generation and provide a structured data analysis:
            
            User Request: {query}
            
            Please provide:
            1. Data Structure: What columns/fields are needed
            2. Data Types: What type of data each column should contain
            3. Sample Size: How many rows of sample data should be generated
            4. Relationships: Any relationships between data fields
            5. Categories: Any categorical data or groupings needed
            6. Time Series: If time-based data is needed, specify the time range and frequency
            
            Format your response as JSON with the following structure:
            {{
                "columns": [
                    {{"name": "column_name", "type": "data_type", "description": "purpose"}},
                    ...
                ],
                "sample_size": number,
                "relationships": ["description of relationships"],
                "categories": ["list of categories if applicable"],
                "time_series": {{"enabled": boolean, "range": "description", "frequency": "daily/weekly/monthly"}},
                "special_requirements": ["any special data requirements"]
            }}
            """
            
            response = model.generate_content(analysis_prompt)
            return response.text
            
        except Exception as e:
            return f"Error in data analysis: {str(e)}"


class ExcelStructureTool(BaseTool):
    """Tool for designing Excel workbook structure"""
    
    name: str = "excel_structure_tool"
    description: str = "Designs the overall structure of Excel workbook including sheets, layouts, and organization"
    
    def _run(self, data_analysis: str, user_request: str) -> str:
        """Design Excel structure based on data analysis"""
        try:
            import google.generativeai as genai
            model = genai.GenerativeModel('gemini-2.0-flash')
            
            structure_prompt = f"""
            Based on this data analysis and user request, design an Excel workbook structure:
            
            Data Analysis: {data_analysis}
            User Request: {user_request}
            
            Please provide a comprehensive Excel structure design including:
            1. Sheet Names: What sheets should be created and their purpose
            2. Sheet Layout: How data should be organized in each sheet
            3. Header Design: Column headers and their formatting
            4. Data Organization: How to group and organize the data
            5. Summary Sheets: Any dashboard or summary sheets needed
            6. Cross-references: Links between sheets if applicable
            
            Format your response as JSON:
            {{
                "sheets": [
                    {{
                        "name": "sheet_name",
                        "purpose": "description",
                        "columns": ["col1", "col2", ...],
                        "layout_type": "data/dashboard/summary",
                        "special_features": ["pivot_table", "charts", "formulas"]
                    }},
                    ...
                ],
                "workbook_structure": {{
                    "main_data_sheet": "sheet_name",
                    "summary_sheets": ["sheet1", "sheet2"],
                    "chart_sheets": ["sheet1", "sheet2"]
                }},
                "cross_references": ["description of links between sheets"]
            }}
            """
            
            response = model.generate_content(structure_prompt)
            return response.text
            
        except Exception as e:
            return f"Error in structure design: {str(e)}"


class ChartDesignTool(BaseTool):
    """Tool for designing charts and visualizations"""
    
    name: str = "chart_design_tool"
    description: str = "Designs appropriate charts and visualizations based on data and user requirements"
    
    def _run(self, data_analysis: str, excel_structure: str, user_request: str) -> str:
        """Design charts for the Excel workbook"""
        try:
            import google.generativeai as genai
            model = genai.GenerativeModel('gemini-2.0-flash')
            
            chart_prompt = f"""
            Design charts and visualizations for this Excel workbook:
            
            Data Analysis: {data_analysis}
            Excel Structure: {excel_structure}
            User Request: {user_request}
            
            Please specify:
            1. Chart Types: What types of charts are most appropriate
            2. Chart Data: What data each chart should visualize
            3. Chart Placement: Which sheet each chart should go on
            4. Chart Styling: Colors, themes, and formatting
            5. Interactive Elements: Any dynamic or interactive features
            
            Available chart types: bar, line, pie, doughnut, column, area, scatter, radar
            
            Format as JSON:
            {{
                "charts": [
                    {{
                        "type": "chart_type",
                        "title": "Chart Title",
                        "sheet": "target_sheet",
                        "data_range": "A1:D10",
                        "position": "E2",
                        "styling": {{
                            "theme": "colorful/monochromatic/custom",
                            "colors": ["color1", "color2", ...],
                            "style": "modern/classic/minimal"
                        }},
                        "features": ["data_labels", "legend", "trendline"]
                    }},
                    ...
                ],
                "dashboard_layout": {{
                    "enabled": boolean,
                    "sheet_name": "Dashboard",
                    "chart_arrangement": "grid/flow/custom"
                }}
            }}
            """
            
            response = model.generate_content(chart_prompt)
            return response.text
            
        except Exception as e:
            return f"Error in chart design: {str(e)}"


class ExcelFormattingTool(BaseTool):
    """Tool for advanced Excel formatting and styling"""
    
    name: str = "excel_formatting_tool"
    description: str = "Applies advanced formatting, conditional formatting, and styling to Excel workbooks"
    
    def _run(self, excel_structure: str, chart_design: str, user_request: str) -> str:
        """Design formatting for the Excel workbook"""
        try:
            import google.generativeai as genai
            model = genai.GenerativeModel('gemini-2.0-flash')
            
            formatting_prompt = f"""
            Design comprehensive formatting for this Excel workbook:
            
            Excel Structure: {excel_structure}
            Chart Design: {chart_design}
            User Request: {user_request}
            
            Please specify:
            1. Color Scheme: Overall color theme for the workbook
            2. Header Formatting: Styling for column headers
            3. Data Formatting: Number formats, date formats, text alignment
            4. Conditional Formatting: Rules for highlighting important data
            5. Professional Styling: Borders, shading, fonts for professional appearance
            6. Print Settings: Page setup, margins, headers/footers
            
            Format as JSON:
            {{
                "color_scheme": {{
                    "primary": "#color",
                    "secondary": "#color",
                    "accent": "#color",
                    "neutral": "#color"
                }},
                "header_formatting": {{
                    "font": "font_name",
                    "size": font_size,
                    "bold": boolean,
                    "background": "#color",
                    "text_color": "#color"
                }},
                "data_formatting": {{
                    "numbers": "format_string",
                    "dates": "format_string",
                    "percentages": "format_string",
                    "currency": "format_string"
                }},
                "conditional_formatting": [
                    {{
                        "rule": "description",
                        "range": "A1:Z100",
                        "condition": "condition_type",
                        "format": "formatting_details"
                    }},
                    ...
                ],
                "professional_styling": {{
                    "borders": "style",
                    "alternating_rows": boolean,
                    "freeze_panes": "A2",
                    "column_widths": "auto/custom"
                }}
            }}
            """
            
            response = model.generate_content(formatting_prompt)
            return response.text
            
        except Exception as e:
            return f"Error in formatting design: {str(e)}"


class QualityAssuranceTool(BaseTool):
    """Tool for quality assurance and validation"""
    
    name: str = "quality_assurance_tool" 
    description: str = "Performs quality checks and validation on the generated Excel workbook"
    
    def _run(self, excel_file_path: str, original_request: str) -> str:
        """Perform quality assurance on the generated Excel file"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            
            # Load and analyze the Excel file
            wb = load_workbook(excel_file_path)
            sheet_names = wb.sheetnames
            
            quality_report = {
                "file_exists": True,
                "sheet_count": len(sheet_names),
                "sheets": sheet_names,
                "data_validation": [],
                "chart_validation": [],
                "formatting_validation": [],
                "overall_score": 0
            }
            
            # Check each sheet
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                
                # Data validation
                data_check = {
                    "sheet": sheet_name,
                    "has_data": sheet.max_row > 1,
                    "has_headers": bool(sheet['A1'].value),
                    "data_rows": sheet.max_row - 1,
                    "data_columns": sheet.max_column
                }
                quality_report["data_validation"].append(data_check)
                
                # Check for charts
                if hasattr(sheet, '_charts'):
                    chart_count = len(sheet._charts)
                    quality_report["chart_validation"].append({
                        "sheet": sheet_name,
                        "chart_count": chart_count
                    })
            
            # Calculate overall score
            score = 0
            if quality_report["sheet_count"] > 0:
                score += 20
            if any(d["has_data"] for d in quality_report["data_validation"]):
                score += 30
            if any(d["has_headers"] for d in quality_report["data_validation"]):
                score += 20
            if quality_report["chart_validation"]:
                score += 30
            
            quality_report["overall_score"] = score
            
            return json.dumps(quality_report, indent=2)
            
        except Exception as e:
            return f"Error in quality assurance: {str(e)}"


# Excel Generation Crew
class ExcelGeneratorCrew:
    """CrewAI-based Excel Generation System"""
    
    def __init__(self):
        """Initialize the Excel generation crew"""
        self.llm = GeminiLLM()
        self.setup_tools()
        self.setup_agents()
    
    def setup_tools(self):
        """Initialize all tools"""
        self.data_analysis_tool = DataAnalysisTool()
        self.excel_structure_tool = ExcelStructureTool()
        self.chart_design_tool = ChartDesignTool()
        self.excel_formatting_tool = ExcelFormattingTool()
        self.quality_assurance_tool = QualityAssuranceTool()
    
    def setup_agents(self):
        """Initialize all agents with their roles and capabilities"""
        
        # Data Analyst Agent
        self.data_analyst = Agent(
            role='Data Analyst',
            goal='Analyze user requirements and design optimal data structures for Excel generation',
            backstory="""You are an expert data analyst with extensive experience in understanding 
            business requirements and translating them into structured data formats. You excel at 
            identifying the most appropriate data types, relationships, and organizational structures 
            for any given scenario.""",
            verbose=True,
            allow_delegation=False,
            llm=self.llm,
            tools=[self.data_analysis_tool]
        )
        
        # Excel Specialist Agent
        self.excel_specialist = Agent(
            role='Excel Specialist',
            goal='Design comprehensive Excel workbook structures with multiple sheets, formulas, and advanced features',
            backstory="""You are a Microsoft Excel expert with deep knowledge of advanced Excel features, 
            formulas, pivot tables, and workbook organization. You can design complex, professional 
            Excel workbooks that are both functional and user-friendly.""",
            verbose=True,
            allow_delegation=False,
            llm=self.llm,
            tools=[self.excel_structure_tool]
        )
        
        # Chart Designer Agent  
        self.chart_designer = Agent(
            role='Chart Designer',
            goal='Create compelling and appropriate visualizations that effectively communicate data insights',
            backstory="""You are a data visualization expert who understands how to choose the right 
            chart types for different data scenarios. You excel at creating visually appealing and 
            informative charts that help users understand their data at a glance.""",
            verbose=True,
            allow_delegation=False,
            llm=self.llm,
            tools=[self.chart_design_tool]
        )
        
        # Quality Assurance Agent
        self.qa_agent = Agent(
            role='Quality Assurance Specialist',
            goal='Ensure the generated Excel files meet high standards of quality, accuracy, and usability',
            backstory="""You are a meticulous quality assurance specialist with expertise in testing 
            and validating Excel workbooks. You ensure that all generated files are error-free, 
            well-formatted, and meet the original requirements.""",
            verbose=True,
            allow_delegation=False,
            llm=self.llm,
            tools=[self.quality_assurance_tool, self.excel_formatting_tool]
        )
    
    def create_excel_generation_tasks(self, user_request: str) -> List[Task]:
        """Create a sequence of tasks for Excel generation"""
        
        # Task 1: Data Analysis
        data_analysis_task = Task(
            description=f"""
            Analyze the following user request for Excel generation:
            
            "{user_request}"
            
            Determine:
            1. What type of data structure is needed
            2. What columns and data types are required
            3. How much sample data should be generated
            4. Any special relationships or categories in the data
            5. Whether time-series data is needed
            
            Provide a detailed analysis that will guide the Excel structure design.
            """,
            agent=self.data_analyst,
            expected_output="A comprehensive JSON analysis of data requirements including columns, types, sample size, and relationships"
        )
        
        # Task 2: Excel Structure Design
        excel_structure_task = Task(
            description=f"""
            Based on the data analysis results, design a comprehensive Excel workbook structure for:
            
            "{user_request}"
            
            Create:
            1. Multiple sheets with clear purposes
            2. Organized data layouts
            3. Professional header designs
            4. Cross-references between sheets where appropriate
            5. Summary or dashboard sheets if beneficial
            
            The structure should be professional and user-friendly.
            """,
            agent=self.excel_specialist,
            expected_output="A detailed JSON specification of the Excel workbook structure including sheets, layouts, and organization"
        )
        
        # Task 3: Chart and Visualization Design
        chart_design_task = Task(
            description=f"""
            Design appropriate charts and visualizations for the Excel workbook based on:
            
            Original Request: "{user_request}"
            Data Analysis: {data_analysis_task}
            Excel Structure: {excel_structure_task}
            
            Create:
            1. Appropriate chart types for the data
            2. Professional styling and color schemes
            3. Proper chart placement within the workbook
            4. Dashboard layouts if multiple charts are needed
            5. Interactive elements where beneficial
            """,
            agent=self.chart_designer,
            expected_output="A comprehensive JSON specification of charts, visualizations, and dashboard layouts"
        )
        
        # Task 4: Quality Assurance and Final Formatting
        qa_task = Task(
            description=f"""
            Design the final formatting and prepare quality assurance criteria for:
            
            Original Request: "{user_request}"
            Excel Structure: {excel_structure_task}
            Chart Design: {chart_design_task}
            
            Ensure:
            1. Professional color schemes and formatting
            2. Appropriate conditional formatting rules
            3. Proper number and date formats
            4. Print-ready layouts
            5. Quality validation criteria
            """,
            agent=self.qa_agent,
            expected_output="A comprehensive JSON specification of formatting rules and quality assurance criteria"
        )
        
        return [data_analysis_task, excel_structure_task, chart_design_task, qa_task]
    
    def generate_excel_with_crew(self, user_request: str, user_id: Optional[str] = None) -> Dict[str, Any]:
        """Generate Excel file using the crew system"""
        try:
            # Create tasks
            tasks = self.create_excel_generation_tasks(user_request)
            
            # Create and execute crew
            crew = Crew(
                agents=[self.data_analyst, self.excel_specialist, self.chart_designer, self.qa_agent],
                tasks=tasks,
                process=Process.sequential,
                verbose=True
            )
            
            # Execute the crew
            crew_result = crew.kickoff()
            
            # Parse the results from each task
            try:
                # Extract results from the crew output
                if hasattr(crew_result, 'tasks_outputs') and crew_result.tasks_outputs:
                    task_outputs = crew_result.tasks_outputs
                    data_analysis = str(task_outputs[0]) if len(task_outputs) > 0 else "{}"
                    excel_structure = str(task_outputs[1]) if len(task_outputs) > 1 else "{}"
                    chart_design = str(task_outputs[2]) if len(task_outputs) > 2 else "{}"
                    formatting_spec = str(task_outputs[3]) if len(task_outputs) > 3 else "{}"
                else:
                    # Fallback: use the crew result as a single output
                    crew_output = str(crew_result)
                    data_analysis = excel_structure = chart_design = formatting_spec = crew_output
                
            except Exception as parse_error:
                print(f"Error parsing crew results: {parse_error}")
                # Use fallback values
                data_analysis = json.dumps({"columns": [{"name": "Sample Data", "type": "text"}], "sample_size": 100})
                excel_structure = json.dumps({"sheets": [{"name": "Data", "purpose": "Main data sheet"}]})
                chart_design = json.dumps({"charts": []})
                formatting_spec = json.dumps({"color_scheme": {"primary": "#4472C4"}})
            
            # Now use the existing Excel generation system with the crew's specifications
            excel_file_path = self.execute_excel_generation(
                user_request, data_analysis, excel_structure, chart_design, formatting_spec
            )
            
            # Perform quality assurance
            qa_result = self.quality_assurance_tool._run(excel_file_path, user_request)
            
            return {
                "success": True,
                "file_path": excel_file_path,
                "crew_analysis": {
                    "data_analysis": data_analysis,
                    "excel_structure": excel_structure,
                    "chart_design": chart_design,
                    "formatting_spec": formatting_spec
                },
                "quality_report": qa_result,
                "process_type": "crew_generated"
            }
            
        except Exception as e:
            print(f"Error in crew-based Excel generation: {e}")
            # Fallback to the original Excel generation system
            return self.fallback_excel_generation(user_request, user_id)
    
    def execute_excel_generation(self, user_request: str, data_analysis: str, 
                                excel_structure: str, chart_design: str, 
                                formatting_spec: str) -> str:
        """Execute the actual Excel file generation based on crew specifications"""
        try:
            # Create enhanced prompt for the existing Excel generator
            enhanced_prompt = f"""
            Generate an Excel file based on these detailed specifications:
            
            Original Request: {user_request}
            
            Data Analysis: {data_analysis}
            Excel Structure: {excel_structure}
            Chart Design: {chart_design}
            Formatting Specifications: {formatting_spec}
            
            Create a professional Excel workbook that implements all these specifications.
            """
            
            # Use the existing Excel generation function with enhanced prompt
            result = generate_excel_from_prompt(enhanced_prompt)
            
            if result.get("success") and result.get("file_path"):
                return result["file_path"]
            else:
                raise Exception(f"Excel generation failed: {result.get('error', 'Unknown error')}")
                
        except Exception as e:
            print(f"Error in Excel generation execution: {e}")
            # Use basic Excel generation as fallback
            result = generate_excel_from_prompt(user_request)
            if result.get("success") and result.get("file_path"):
                return result["file_path"]
            else:
                raise Exception(f"Fallback Excel generation also failed: {result.get('error', 'Unknown error')}")
    
    def fallback_excel_generation(self, user_request: str, user_id: Optional[str] = None) -> Dict[str, Any]:
        """Fallback to original Excel generation system if crew fails"""
        try:
            result = generate_excel_from_prompt(user_request)
            if result.get("success"):
                return {
                    "success": True,
                    "file_path": result["file_path"],
                    "process_type": "fallback_generated",
                    "crew_analysis": None,
                    "quality_report": "Generated using fallback system"
                }
            else:
                return {
                    "success": False,
                    "error": result.get("error", "Excel generation failed"),
                    "process_type": "fallback_failed"
                }
        except Exception as e:
            return {
                "success": False,
                "error": f"Fallback generation failed: {str(e)}",
                "process_type": "complete_failure"
            }


# Global crew instance
_excel_crew = None

def get_excel_crew() -> ExcelGeneratorCrew:
    """Get or create the global Excel generation crew instance"""
    global _excel_crew
    if _excel_crew is None:
        _excel_crew = ExcelGeneratorCrew()
    return _excel_crew

def generate_excel_with_crew(user_request: str, user_id: Optional[str] = None) -> Dict[str, Any]:
    """Public function to generate Excel files using the crew system"""
    crew = get_excel_crew()
    return crew.generate_excel_with_crew(user_request, user_id)
